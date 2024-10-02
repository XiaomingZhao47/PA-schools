import openpyxl
from utils import Logger, detect_year
import shutil
import os
from xls2xlsx import XLS2XLSX
import re

class SheetDict:
    def __init__(self, dict, identifier):
        self.dict = dict
        self.identifier = identifier

def do_conversions(ORGANIZED_DATA_DIRECTORY, logger):
    logger.write("Converting files...")
    logger.indent()

    for subdirectory in os.listdir(ORGANIZED_DATA_DIRECTORY):
        for filename in os.listdir(ORGANIZED_DATA_DIRECTORY + "/" + subdirectory):
            file = ORGANIZED_DATA_DIRECTORY + "/" + subdirectory + "/" + filename
            logger.write(f'Processing {file}')
            logger.indent()

            if ".xls" not in file:
                logger.write("Invalid file")
                logger.unindent()
                continue

            if ".xlsx" not in file:
                logger.write(f'xls detected!')

                xlsx_file = ".".join(xls_file.split(".")[0:-1]) + ".xlsx"
                XLS2XLSX(xls_file).to_xlsx(xlsx_file)

                os.remove(xls_file)

            logger.unindent()

    logger.unindent()
    logger.write("Done!")

def detect_type(value):
    if not isinstance(value, str):
        return value

    new_value = value.strip()

    if new_value.isdigit():
        return int(new_value)

    # Matches floating point value
    if bool(re.match(r'^[+-]?(\d+(\.\d*)?|\.\d+)$', new_value)):
        return float(new_value)

    # Matches phone number
    if bool(re.match(r'\d{3}\-\d{3}\-\d{4}', new_value)):
        return int(new_value[0:3] + new_value[4:7] + new_value[8:12])

    return new_value

def rename_attribute(attribute):
    if attribute == None:
        return None

    new_name = attribute.lower()
    new_name = "_".join(new_name.split())
    new_name = new_name.replace(",", "")
    new_name = new_name.replace("&", "and")
    new_name = new_name.replace("district_name", "lea_name")

    if new_name == "school_district":
        return "lea_name"

    return new_name


def rename_afr_exp_attribute(attribute):
    new_name = rename_attribute(attribute)
    if new_name is None:
        return None
    if "actual_instruction_expense" in new_name:
        return None

    search = re.search(r'\d{4}$', new_name)

    if search is None:

        if new_name == "aun":
            return None
        if "cuurent" in new_name:
            return None
        if "current" in new_name:
            return None
        if new_name == "ctgy":
            return None
        if new_name == "cat":
            return None
        if "actual_instruction" in new_name:
            return None
        if new_name == "school_district":
            return "lea_name"

        return new_name

    id = int(search.group(0))
    match id:
        case 1000:
            return "instruction"
        case 1100:
            return "regular_programs"
        case 1200:
            return "gifted_programs"
        case 1300:
            return "vocational_programs"
        case 1400:
            return "other_instructional"
        case 1500:
            return "nonpublic_programs"
        case 1600:
            return "adult_programs"
        case 1700:
            return "secondary_programs"
        case 1800:
            return "pre-k"
        case 2000:
            return "support_services"
        case 2100:
            return "personell"
        case 2200:
            return "staff"
        case 2300:
            return "administration"
        case 2400:
            return "health"
        case 2500:
            return "business"
        case 2600:
            return "plant_ops"
        case 2700:
            return "transportation"
        case 2800:
            return "central"
        case 2900:
            return "other_support"
        case 3000:
            return "non-instructional_services"
        case 4000:
            return "facai"
        case 5000:
            return "oefu"
        case _:
            print("Invalid ID!")
            print(id)
            print(new_name)
            exit()

def rename_afr_exp_adm_attr(attribute):
    new_name = rename_attribute(attribute)
    if new_name is None:
        return None
    if new_name == "aun":
        return None
    if new_name == "ctgy":
            return None
    if new_name == "cat":
        return None

    if bool(re.match(r'^\d{4}\-\d{2}\_', new_name)):
        new_name = new_name[8:]

    new_name = new_name.replace("memebership", "membership")

    return new_name

def rename_fast_fact_attribute(attribute):
    new_name = rename_attribute(attribute)

    new_name = new_name.replace("_-_percent_enrollment_by_student_groups", "")
    new_name = new_name.replace("_-_percent_enrollment_by_race/ethnicity", "")

    if "2_or_more_races" in new_name:
        return new_name.replace("2_or_more_races", "multiracial")

    if "american_indian/alaskan_native" in new_name:
        return "ai_an"

    if "black/african_american" in new_name:
        return "african_american"

    if "career_and_technical_center" in new_name:
        return new_name.replace("career_and_technical_center", "ctc")

    if "native_hawaiian_or_other_pacific_islander" in new_name:
        return "nh_pi"

    if "intermediate_unit" in new_name:
        return new_name.replace("intermediate_unit", "iu")

    if "geographic_size" in new_name:
        return "district_size_sq_mi"

    if "male_(school)" in new_name:
        return "male"

    if "female_(school)" in new_name:
        return "female"

    return new_name

def parse_district_fast_facts(wb):
    districts = {}
    sheet = wb.active

    first = True
    for row in sheet.rows:
        if first:
            first = False
            continue

        district_name = row[0].value
        aun = row[1].value
        attribute = rename_fast_fact_attribute(row[2].value)
        value = detect_type(row[3].value)

        if "offered" in attribute:
            continue

        if aun not in districts:
            districts[aun] = {"lea_name": district_name}

        districts[aun][attribute] = value

    return SheetDict(districts, "aun")

def parse_school_fast_facts(wb):

    schools = {}
    sheet = wb.active

    first = True
    for row in sheet.rows:
        if first:
            first = False
            continue

        district_name = row[0].value
        school_name = row[1].value
        aun = row[2].value
        school_id = row[3].value
        attribute = rename_fast_fact_attribute(row[4].value)
        value = row[5].value

        if "district_1" in attribute or "district_2" in attribute:
            continue

        if school_id not in schools:
            schools[school_id] = {"school_name": school_name, "district_name": district_name, "aun": aun}

        schools[school_id][attribute] = detect_type(value)

    return SheetDict(schools, "school_id")

def parse_afr_expenditure(wb, year):
    expenditures = {}
    sheet = wb.active

    first = True
    for rowIdx, row in enumerate(sheet.rows):
        if first:
            first = False
            continue

        aun = row[1].value

        if aun is None:
            continue

        expenditures[aun] = {}

        for colIdx, col in enumerate(row):

            attribute = sheet.cell(row=1, column=colIdx+1).value
            value = sheet.cell(row=rowIdx+1, column=colIdx+1).value

            attribute = rename_afr_exp_attribute(attribute)
            if attribute is None:
                continue

            expenditures[aun][attribute] = detect_type(value)

    expenditures_sheet =  SheetDict(expenditures, "aun")

    expenditures_per_adm = {}
    sheet = wb.worksheets[1]

    first = True
    for rowIdx, row in enumerate(sheet.rows):
        if first:
            first = False
            continue

        if year == 2018 or year == 2022:
            aun = row[1].value
        else:
            aun = row[0].value

        if aun is None:
            continue

        expenditures_per_adm[aun] = {}

        for colIdx, col in enumerate(row):

            attribute = sheet.cell(row=1, column=colIdx+1).value
            value = sheet.cell(row=rowIdx+1, column=colIdx+1).value

            attribute = rename_afr_exp_adm_attr(attribute)
            if attribute is None:
                continue

            expenditures_per_adm[aun][attribute] = detect_type(value)
    expenditures_per_adm_sheet =  SheetDict(expenditures_per_adm, "aun")

    return [expenditures_sheet, expenditures_per_adm_sheet]
def write_dicts(classified_sheet_dicts, CLEAN_DATA_DIRECTORY):
    for classification, sheet_dicts in classified_sheet_dicts.items():

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        next_key_index = 2
        key_indices = {}

        for year, sheet_dict in dict(sorted(sheet_dicts.items())).items():

            sheet_name = str(year)
            sheet = wb.create_sheet(title=sheet_name)

            rowIdx = 2

            sheet.cell(row=1, column=1).value = sheet_dict.identifier

            for key, record in sheet_dict.dict.items():
                sheet.cell(row=rowIdx, column=1).value = key

                for record_key, attribute in record.items():
                    if record_key not in key_indices:
                        key_indices[record_key] = next_key_index
                        next_key_index = next_key_index + 1


                    sheet.cell(row=1, column = key_indices[record_key]).value = record_key

                    sheet.cell(row=rowIdx, column=key_indices[record_key]).value = attribute
                rowIdx = rowIdx + 1

        wb.save(CLEAN_DATA_DIRECTORY + "/" + classification + ".xlsx")

def clean_data(ORGANIZED_DATA_DIRECTORY, CLEAN_DATA_DIRECTORY, logger):
    logger.write("Cleaning Data...")
    logger.indent()

    for subdirectory in os.listdir(ORGANIZED_DATA_DIRECTORY):
        sheet_dicts = {}

        if "Fast_Facts" not in subdirectory and "AFR_Expenditure" not in subdirectory:
            continue

        for filename in os.listdir(ORGANIZED_DATA_DIRECTORY + "/" + subdirectory):
            if "#" in filename:
                continue

            logger.write(filename)
            file = ORGANIZED_DATA_DIRECTORY + "/" + subdirectory + "/" + filename
            new_file = CLEAN_DATA_DIRECTORY + "/" + filename

            year = detect_year(file)
            wb = openpyxl.open(file)

            if "Fast_Facts_District" in file:
                if "Fast_Facts_District" not in sheet_dicts:
                    sheet_dicts["Fast_Facts_District"] = {}
                sheet_dicts["Fast_Facts_District"][year] = parse_district_fast_facts(wb)

            elif "Fast_Facts_School" in file:
                if "Fast_Facts_School" not in sheet_dicts:
                    sheet_dicts["Fast_Facts_School"] = {}
                sheet_dicts["Fast_Facts_School"][year] = parse_school_fast_facts(wb)

            elif "AFR_Expenditure" in file:
                if "AFR_Expenditure" not in sheet_dicts:
                    sheet_dicts["AFR_Expenditure"] = {}
                    sheet_dicts["AFR_Expenditure_Per_ADM"] = {}

                afr_expenditures = parse_afr_expenditure(wb, year)
                sheet_dicts["AFR_Expenditure"][year] = afr_expenditures[0]
                sheet_dicts["AFR_Expenditure_Per_ADM"][year] = afr_expenditures[1]
            else:
                logger.write(f'No parser for: {file}')
                continue

        write_dicts(sheet_dicts, CLEAN_DATA_DIRECTORY)

    logger.unindent()
    logger.write("Done!")

def remove_extra_files(ORGANIZED_DATA_DIRECTORY, logger):
    logger.write("Removing Extra Files...")
    logger.indent()

    for subdirectory in os.listdir(ORGANIZED_DATA_DIRECTORY):
        for filename in os.listdir(ORGANIZED_DATA_DIRECTORY + "/" + subdirectory):
            file = ORGANIZED_DATA_DIRECTORY + "/" + subdirectory + "/" + filename
            logger.write(f'Processing {file}')
            logger.indent()

            if ".xlsx" not in file:
                os.remove(file)
                logger.write("Removing file")

            logger.unindent()


    logger.unindent()
    logger.write("Done!")

def run(ORGANIZED_DATA_DIRECTORY, CLEAN_DATA_DIRECTORY, logger):
    logger.indent()

    do_conversions(ORGANIZED_DATA_DIRECTORY, logger)
    remove_extra_files(ORGANIZED_DATA_DIRECTORY, logger)

    clean_data(ORGANIZED_DATA_DIRECTORY, CLEAN_DATA_DIRECTORY, logger)

    logger.unindent()

#logger = Logger("crawler-logs.txt")
#logger.write("Staring Script...")
#run("./data-organized", "./data-clean", logger)
#logger.write("Done!")
