import openpyxl
from utils import Logger, detect_year, detect_type, SheetDict
import shutil
import os
from xls2xlsx import XLS2XLSX
import re

schools = SheetDict({}, "school_id")
leas = SheetDict({}, "aun")
ius = SheetDict({}, "aun")

def can_safely_replace(val1, val2):
    if val1 == val2:
        return True

    if not (type(val1) == str and type(val2) == str):
        return False

    val1 = val1.lower().replace(" ", "")
    val2 = val2.lower().replace(" ", "")

    val1 = val1.replace("charterschool", "cs")
    val2 = val2.replace("charterschool", "cs")

    val1 = val1.replace("saint", "st.")
    val2 = val2.replace("saint", "st.")

    return val1 == val2


def add_to_sheet_dict(logger, sheet_dict, id, attribute, value):
    dict = sheet_dict.dict

    if id not in dict:
        dict[id] = {}

    if attribute in dict[id]:
        old_val = dict[id][attribute]
        if not can_safely_replace(old_val, value):
            logger.warn(f'Clobbering {attribute} in dict[{id}]. Replacing {old_val} with {value}');

    dict[id][attribute] = value

def add_to_composite_dict(logger, composite_dict, id, year, attribute, value):
    if year not in composite_dict:
        composite_dict[year] = {}

    if id not in composite_dict[year]:
        composite_dict[year][id] = {}

    if attribute in composite_dict[year][id]:
        old_val = composite_dict[year][id][attribute]
        if not can_safely_replace(old_val, value):
            logger.warn(f'Clobbering {attribute} in composite_dict[{year}][{id}]. Replacing {old_val} with {value}');

    composite_dict[year][id][attribute] = value

class SheetDict:
    def __init__(self, dict, identifier):
        self.dict = dict
        self.identifier = identifier

def parse_wb(wb, logger):
    data_dict = {}

    for sheet in wb.worksheets:
        year = detect_type(sheet.title)

        for col_idx, col in enumerate(sheet.iter_cols()):
            if col_idx == 0:
                continue

            attr = col[0].value

            is_lea_attr = attr in ["lea_name", "county", "district_address_(street)", "district_address_(city)", "district_address_(state)", "district_zip_code", "website", "telephone_number"]
            is_iu_attr = attr in ["iu_name"]

            for row_idx, cell in enumerate(col):
                if row_idx == 0:
                    continue

                aun = sheet.cell(row=row_idx+1, column=1).value
                value = cell.value

                if is_lea_attr:
                    add_to_sheet_dict(logger, leas, aun, attr, value)
                elif is_iu_attr:
                    add_to_sheet_dict(logger, ius, aun, attr, value)
                else:
                    add_to_composite_dict(logger, data_dict, aun, year, attr, value)

    return SheetDict(data_dict, "aun")

def write_composite_dict(sheet_dict, filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = sheet_dict.identifier
    sheet.cell(row=1, column=2).value = "year"

    next_key_index = 3
    key_indices = {}

    rowIdx = 2

    for year, year_dict in sheet_dict.dict.items():
        for id, record in year_dict.items():
            for record_key, attribute in record.items():
                sheet.cell(row=rowIdx, column=1).value = id
                sheet.cell(row=rowIdx, column=2).value = year

                if record_key not in key_indices:
                    key_indices[record_key] = next_key_index
                    next_key_index = next_key_index + 1

                sheet.cell(row=1, column = key_indices[record_key]).value = record_key
                sheet.cell(row=rowIdx, column=key_indices[record_key]).value = attribute
            rowIdx = rowIdx + 1

    wb.save(filename)

def write_sheet_dict(sheet_dict, filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = sheet_dict.identifier

    next_key_index = 2
    key_indices = {}

    rowIdx = 2

    for id, record in sheet_dict.dict.items():
        for record_key, attribute in record.items():
            sheet.cell(row=rowIdx, column=1).value = id

            if record_key not in key_indices:
                key_indices[record_key] = next_key_index
                next_key_index = next_key_index + 1

            sheet.cell(row=1, column = key_indices[record_key]).value = record_key
            sheet.cell(row=rowIdx, column=key_indices[record_key]).value = attribute
        rowIdx = rowIdx + 1


    wb.save(filename)

def run(CLEAN_DATA_DIRECTORY, NORMALIZED_DATA_DIRECTORY, logger):
    logger.indent()

    print(f'Clean Data: {CLEAN_DATA_DIRECTORY}')

    for filename in os.listdir(CLEAN_DATA_DIRECTORY):
        if "#" in filename:
            continue

        logger.write(f'Processing {filename}')

        if "AFR" not in filename and "IU" not in filename and "Fast_Facts_District" not in filename:
            continue

        file = CLEAN_DATA_DIRECTORY + "/" + filename
        new_file = NORMALIZED_DATA_DIRECTORY + "/" + filename
        wb = openpyxl.open(file)

        dict = parse_wb(wb, logger)
        write_composite_dict(dict, new_file)


    write_sheet_dict(schools, NORMALIZED_DATA_DIRECTORY + "/Schools.xlsx")
    write_sheet_dict(leas, NORMALIZED_DATA_DIRECTORY + "/Leas.xlsx")
    write_sheet_dict(ius, NORMALIZED_DATA_DIRECTORY + "/IUs.xlsx")




    logger.unindent()

#logger = Logger("crawler-logs.txt")
#logger.write("Staring Script...")
#run("./data-organized", "./data-clean", logger)
#logger.write("Done!")
