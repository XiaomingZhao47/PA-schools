'''
<FILE>
clean_data.py


<DESCRIPTION>
The purpose of this script is to clean the data files. It does so through a few
important steps:

    * Ensure File Type Consistencty
        All files, if possible, will be converted to an .xslx format. Any files
        that cannot be converted will be removed. This step ensures that we can
        use the same system to process all files.

    * Clean Files:
        Once all the files are converted to an .xslx format, this step will
        remove any unecessary file formatting, and restructure the data fields
        in each file to follow a consistent format. It will also merge each
        statistic type's years into one file. (See @clean_files(...) for more
        details).

<FUNCTIONS>
This script can be run by calling clean_data.run(<args>). All other functions
in this script should remain private. This section only lists a brief description
of each function. For more comprehensive documentation, see each method directly.

    * run(...):
        Converts, removes, and cleans the data files.

    * do_conversions(...):
        Converts each file, if possible, to an .xlsx format.

    * remove_extra_files(...)
        Removes any files not in an .xlsx format.

    * clean_files(...):
        Processes each file to follow a consistent structure.

    * trim_sheet(...)
        Removes leading rows and columns from a sheet.

    * parse_standard_sheet(...)
        Parses a standard data sheet.

    * rename_attribute(...)
        Renames attributes to follow a consistent and safe naming convention.

    * rename_fast_fact_attribute(...):
        Renames School/District Fast Fact attributes.

    * parse_district_fast_facts(...):
        Parses District Fast Fact worksheets.

    * parse_school_fast_facts(...)
        Parses School Fast Fact worksheets.

    * parse_afr_expenditure(...)
        Parses AFR Expenditure worksheets.

    * parse_afr_revenue(...)
        Parses AFR Revenute worksheets.

    * parse_aid_ratio(...)
        Parses Aid Ratio worksheets.

    * parse_keystone(...)
        Parses Keystone Exam worksheets.

    * parse_apd(...)
        Parses APD worksheets.

    * parse_cohort(...)
        Parses Cohort worksheets.

    * write_dicts(...)
        Writes the SheetDicts to file.
'''

import openpyxl
import shutil
import os
import re
from xls2xlsx import XLS2XLSX
from pathlib import Path
from scripts.utils import Logger, detect_year, SheetDict, detect_type

def run(ORGANIZED_DATA_DIRECTORY, CLEAN_DATA_DIRECTORY, logger):
    '''
    Converts, removes, and cleans the data files.

    <ARGUMENTS>
        * ORGANIZED_DATA_DIRECTORY [String]: The path to the input directory.

        * CLEAN_DATA_DIRECTORY [String]: The path to the output directory.

        * logger [utils.Logger]: The current Logger instance.
    '''

    logger.indent()

    do_conversions(ORGANIZED_DATA_DIRECTORY, logger)
    remove_extra_files(ORGANIZED_DATA_DIRECTORY, logger)
    clean_files(ORGANIZED_DATA_DIRECTORY, CLEAN_DATA_DIRECTORY, logger)

    logger.unindent()

def do_conversions(ORGANIZED_DATA_DIRECTORY, logger):
    '''
    Removes any files not in an .xlsx format.

    <EXTENDED_DESCRIPTION>
    All modification will be done in-place.

    <ARGUMENTS>
        * ORGANIZED_DATA_DIRECTORY [String]: The path to the input directory.

        * logger [utils.Logger]: The current Logger instance.
    '''

    logger.write("Converting files...")
    logger.indent()

    for subdirectory in os.listdir(ORGANIZED_DATA_DIRECTORY):
        for filename in os.listdir(ORGANIZED_DATA_DIRECTORY + "/" + subdirectory):
            file = ORGANIZED_DATA_DIRECTORY + "/" + subdirectory + "/" + filename
            logger.write(f'Processing {file}')
            logger.indent()

            if ".xls" not in file:
                logger.write("Invalid file")
                logger.unindent()
                continue

            if ".xlsx" not in file:
                logger.write(f'xls detected!')

                xlsx_file = ".".join(file.split(".")[0:-1]) + ".xlsx"
                XLS2XLSX(file).to_xlsx(xlsx_file)

                os.remove(file)

            logger.unindent()

    logger.unindent()
    logger.write("Done!")

def remove_extra_files(ORGANIZED_DATA_DIRECTORY, logger):
    '''
    Converts each file, if possible, to an .xlsx format.

    <EXTENDED_DESCRIPTION>
    All modification will be done in-place.

    <ARGUMENTS>
        * ORGANIZED_DATA_DIRECTORY [String]: The path to the input directory.

        * logger [utils.Logger]: The current Logger instance.
    '''

    logger.write("Removing Extra Files...")
    logger.indent()

    for subdirectory in os.listdir(ORGANIZED_DATA_DIRECTORY):
        for filename in os.listdir(ORGANIZED_DATA_DIRECTORY + "/" + subdirectory):
            file = ORGANIZED_DATA_DIRECTORY + "/" + subdirectory + "/" + filename
            logger.write(f'Processing {file}')
            logger.indent()

            if ".xlsx" not in file:
                os.remove(file)
                logger.write("Removing file")

            logger.unindent()


    logger.unindent()
    logger.write("Done!")

def clean_files(ORGANIZED_DATA_DIRECTORY, CLEAN_DATA_DIRECTORY, logger):
    '''
    Processes each file to follow a consistent structure.

    <EXTENDED_DESCRIPTION>
    To clean the data files, the data is written into a SheetDict. See <utils.py>
    for a description of SheetDicts.

    It reorganizes the files in a few imporant ways:

        * Splits Files Into Sheets:
            Sheets within each file are separated. E.g., if a file,
            "Stat1_2019-2020.xlsx" has sheets "Sheet1", "Sheet2", it will split
            the file into two, named "Sheet1.xlsx" and "Sheet2.xlsx".

        * Combines Sheets:
            When performing the previous step, it will combine the statistics
            per year. E.g., There are three files, named "Stat1_2019-2020.xlsx",
            "Stat1_2020-2021.xslx", and "Stat2_2021-2022.xlsx", each consisting
            of a singular sheet. The output will be two files, "Stat1.xlsx",
            containing sheets [2019, 2020], and "Stat2.xlsx", containing a
            sheet only for 2021.

        * Matching Columns:
            Columns will remain consistant between years, per statistic. E.g.
            if the output file, "Stat1.xlsx', has the attribute "mvpi" in
            column D in year 2019, all other years for in "Stat1.xlsx' will
            have "mvpi" as the Dth column.
    '''

    logger.write("Cleaning Data...")
    logger.indent()

    for subdirectory in os.listdir(ORGANIZED_DATA_DIRECTORY):
        sheet_dicts = {}

        # Don't process files without parsers written yet
        should_process = False
        for process_check in ["Fast", "AFR", "Aid", "Cohort", "Keystone"]:
            if process_check in subdirectory:
                should_process = True

        if not should_process:
            continue

        #if "Fast" not in subdirectory and "AFR" not in subdirectory and "Aid" not in subdirectory and "APD" not in subdirectory and "Cohort" not in subdirectory:
        #if "Keystones" not in subdirectory:
        #    continue

        for filename in os.listdir(ORGANIZED_DATA_DIRECTORY + "/" + subdirectory):
            if "#" in filename:
                continue

            logger.write(filename)
            file = ORGANIZED_DATA_DIRECTORY + "/" + subdirectory + "/" + filename
            new_file = CLEAN_DATA_DIRECTORY + "/" + filename

            year = detect_year(file)
            wb = openpyxl.open(file)

            # 2023 follows a different format than the rest of the years
            if "Fast_Facts_District" in file and year != 2023:
                if "Fast_Facts_District" not in sheet_dicts:
                    sheet_dicts["Fast_Facts_District"] = {}
                sheet_dicts["Fast_Facts_District"][year] = parse_district_fast_facts(wb)

            elif "Fast_Facts_School" in file and year != 2023:
                if "Fast_Facts_School" not in sheet_dicts:
                    sheet_dicts["Fast_Facts_School"] = {}
                sheet_dicts["Fast_Facts_School"][year] = parse_school_fast_facts(wb)

            elif "AFR_Expenditure" in file:
                if "AFR_Expenditure" not in sheet_dicts:
                    sheet_dicts["AFR_Expenditure"] = {}
                    #sheet_dicts["AFR_Expenditure_Per_ADM"] = {}

                afr_expenditures = parse_afr_expenditure(wb, year)
                sheet_dicts["AFR_Expenditure"][year] = afr_expenditures[0]
                #sheet_dicts["AFR_Expenditure_Per_ADM"][year] = afr_expenditures[1] Doesn't provide new data

            elif "AFR_Revenue" in file:
                if "AFR_Revenue" not in sheet_dicts:
                    sheet_dicts["AFR_Revenue"] = {}
                    sheet_dicts["AFR_Revenue_Per_ADM"] = {}
                    sheet_dicts["AFR_Revenue_TCEM"] = {}


                afr_revenues = parse_afr_revenue(wb, year)
                sheet_dicts["AFR_Revenue"][year] = afr_revenues[0]
                sheet_dicts["AFR_Revenue_Per_ADM"][year] = afr_revenues[1] #Doesn't provide new data
                sheet_dicts["AFR_Revenue_TCEM"][year] = afr_revenues[2]

            elif "Aid_Ratios" in file:
                if "Aid_Ratios_LEA" not in sheet_dicts:
                    sheet_dicts["Aid_Ratios_LEA"] = {}
                    sheet_dicts["Aid_Ratios_IU"] = {}

                aid_ratios = parse_aid_ratio(wb, year)
                sheet_dicts["Aid_Ratios_LEA"][year] = aid_ratios[0]
                sheet_dicts["Aid_Ratios_IU"][year] = aid_ratios[1]

            #elif "APD" in file:
            #    if "APD" not in sheet_dicts:
            #        sheet_dicts["APD"] = {}
            #   sheet_dicts["APD"][year] = parse_apd(wb)

            elif "Cohort" in file:
                if "Cohort_LEA" not in sheet_dicts:
                    sheet_dicts["Cohort_LEA"] = {}
                    sheet_dicts["Cohort_School"] = {}

                fix = "Four" in subdirectory and year == 2012

                cohorts = parse_cohort(wb, year, fix)
                sheet_dicts["Cohort_LEA"][year] = cohorts[0]
                sheet_dicts["Cohort_School"][year] = cohorts[1]
            elif "Keystone_Exams_School" in file:
                if "Keystone_Exams" not in sheet_dicts:
                    sheet_dicts["Keystone_Exams"] = {}

                sheet_dicts["Keystone_Exams"][year] = parse_keystone(wb, year)
            else:
                logger.write(f'No parser for: {file}')
                wb.close()
                continue
            wb.close()
        write_dicts(sheet_dicts, subdirectory, CLEAN_DATA_DIRECTORY)

    logger.unindent()
    logger.write("Done!")

def trim_sheet(sheet, rows, cols=0):
    '''
    Removes leading rows and columns from a sheet.

    <ARGUMENTS>
        * sheet [openpyxl Sheet]: The sheet to be trimmed.

        * rows [Integer]: The number of rows to remove.

        * cols=0 [Integer]: The number of columns to remove.

    <RETURN>
        * [openpyxl Sheet]: The trimmed sheet
    '''

    if rows != 0:
        sheet.delete_rows(1, rows)

    if cols != 0:
        sheet.delete_cols(1, cols)

    return sheet

def parse_standard_sheet(sheet, year, rename_attr_cb, get_id_cb):
    '''
    Parses a standard data sheet.

    <ENTENDED_DESCRIPTION>
    Provides a generic implementation to parse an openpyxl Sheet. The exact
    process on how it parses the sheet will be specified using the callback
    methods provided as input.

    A "standard" sheet is one where each Key value is only provided in the sheet
    once, and the sheet doesn't require modifications other than formatting,
    trimming, and renaming attributes.

    <ARGUMENTS>
        * sheet [openpyxl Sheet]:
            The openpyxl Sheet to be parsed.

        * year [Integer]:
            The year which the Sheet describes.

        * rename_attr_cb [Function]:
            Callback function to rename all attributes in the given sheet.

            <ARGUMENTS>:
                * [String]: The name of the attribute to be renamed.

            <RETURN>:
                * [String]: The attribute's new name.

        * get_id_cb [Function]:
            Callback function to determine each row's key

            <EXTENDED_DESCRIPTION>

            Because the id (aun, school_id, etc.), might not appear in the same
            column year-to-year, this callback function is used to locate that
            identifier.

            <ARGUMENTS>:
                * [ [openpyxl Cell...] ]: The array of data values for each row.
                * [Integer]: The year which the Sheet describes.

            <RETURN>:
                * [Any]: The row's id.

    <RETURN>
        * [Dictionary]: A dictionary containing the entire sheet's data.
    '''

    parsed_sheet = {}

    first = True
    for row_idx, row in enumerate(sheet.rows):

        if first:
            first = False
            continue

        id = detect_type(get_id_cb(row, year))
        if id is None:
            continue

        parsed_sheet[id] = {}
        for col_idx, cell in enumerate(row):

            attribute = rename_attr_cb(sheet.cell(row=1, column=col_idx+1).value)
            value = detect_type(sheet.cell(row=row_idx+1, column=col_idx+1).value)

            if attribute is not None:
                parsed_sheet[id][attribute] = (value)

    return parsed_sheet

def rename_attribute(attribute):
    '''
    Renames attributes to follow a consistent and safe naming convention.

    <ENTENDED_DESCRIPTION>
    This function is indented to provide a common renaming functionality,
    applicable to all parsed sheets. Each parser function should extend this
    renaming function, to tailer each to the sheet.

    The processing provided here includes removing punctuation, replacing
    spaces with underlines, and converting all text to lowercase.

    <ARGUMENTS>
        * attribute [String | None]: The name of the attribute to be renamed.

    <RETURN>
        * [String | None]:
            The renamed attribute, should it be one worth keeping.
            None, otherwise.
    '''

    if attribute == None:
        return None

    new_name = attribute.lower()
    new_name = "_".join(new_name.split())
    new_name = new_name.replace(",", "")
    new_name = new_name.replace("§", "")
    new_name = new_name.replace("&", "and")
    new_name = new_name.replace("district_name", "lea_name")

    if new_name == "school_district" or new_name == "district" :
        return "lea_name"
    if new_name == "schl" or new_name == "school_number":
        return "school_id"
    if new_name == "school":
        return "school_name"

    return new_name

def rename_fast_fact_attribute(attribute):
    '''
    Renames School/District Fast Fact attributes.

    <EXTENDED_DESCRIPTION>
    Because both School Fast Fact and District Fast Fact sheets follow a very
    similar naming system, this function is able to rename both. Hence, instead
    of having one copy for each parser, this was put into the outer scope.

    <ARGUMENTS>
        * attribute [String | None]: The name of the attribute to be renamed.

    <RETURN>
        * [String | None]:
            The renamed attribute, should it be one worth keeping.
            None, otherwise.
    '''

    new_name = rename_attribute(attribute)

    new_name = new_name.replace("_-_percent_enrollment_by_student_groups", "").replace("district_", "lea_")
    new_name = new_name.replace("_-_percent_enrollment_by_race/ethnicity", "").replace("district_", "lea")
    new_name = new_name.replace("(", "").replace(")", "")

    if "2_or_more_races" in new_name:
        return new_name.replace("2_or_more_races", "multiracial")

    if "gifted" in new_name:
        return "gifted"

    if "american_indian/alaskan_native" in new_name:
        return "ai_an"

    if "black/african_american" in new_name:
        return "african_american"

    if "english_learner" in new_name:
        return "english_learner"

    if "career_and_technical_center" in new_name:
        return new_name.replace("career_and_technical_center", "ctc")

    if "native_hawaiian_or_other_pacific_islander" in new_name:
        return "nh_pi"

    if "enrollment_in" in new_name:
        return "ctc_enrollment"

    if "charter_school" in new_name:
        return "cs_enrollment"

    if "intermediate_unit" in new_name:
        return new_name.replace("intermediate_unit", "iu")

    if "geographic_size" in new_name:
        return "district_size"

    if "female_(school)" in new_name:
        return "female"

    if "male_(school)" in new_name:
        return "male"

    if "(street)" in new_name:
        return new_name.replace("(street)", "street")

    if "(city)" in new_name:
        return new_name.replace("(city)", "city")

    if  "(state)" in new_name:
        return new_name.replace("(state)", "state")

    if "zip" in new_name:
        return new_name.replace("zip_code", "address_zip")

    if "telephone_number" in new_name:
        return new_name.replace("telephone_number", "telephone")

    return new_name

def parse_district_fast_facts(wb):
    '''
    Parses District Fast Fact worksheets.

    <ENTENDED_DESCRIPTION>
    Because the District Fast Fact worksheets follow an unusual file structure
    (with what should be columns instead being rows), this function cannot rely
    on @parse_standard_sheet(...). Instead, it implements a custom parsing
    algorithm tailored to this unique structure.

    <ARGUMENTS>
        * wb [openpyxl Workbook]: The worksheet to be parsed.

    <RETURN>
        * [utils.SheetDict]: A SheetDict containing the parsed data.
    '''

    def rename_district_fast_facts_attribute(attr):
        new_name = rename_fast_fact_attribute(attr)
        if new_name is None:
            return None
        if new_name == "website":
            return "lea_website"
        if new_name == "telephone":
            return "lea_telephone"

        return new_name

    districts = {}
    sheet = wb.active

    first = True
    for row in sheet.rows:
        if first:
            first = False
            continue

        district_name = row[0].value
        aun = detect_type(row[1].value)
        attr = rename_district_fast_facts_attribute(row[2].value)
        value = detect_type(row[3].value)

        if attr is None:
            continue
        if "offered" in attr:
            continue

        if aun not in districts:
            districts[aun] = {"lea_name": district_name}

        districts[aun][attr] = value

    return SheetDict(districts, "aun")

def parse_school_fast_facts(wb):
    '''
    Parses School Fast Fact worksheets.

    <ENTENDED_DESCRIPTION>
    Because the School Fast Fact worksheets follow an unusual file structure
    (with what should be columns instead being rows), this function cannot rely
    on @parse_standard_sheet(...). Instead, it implements a custom parsing
    algorithm tailored to this unique structure.

    <ARGUMENTS>
        * wb [openpyxl Workbook]: The worksheet to be parsed.

    <RETURN>
        * [utils.SheetDict]: A SheetDict containing the parsed data.
    '''

    def rename_school_fast_facts_attribute(attr):
        new_name = rename_fast_fact_attribute(attr)
        if new_name is None:
            return None
        if new_name == "website":
            return "school_website"
        if new_name == "telephone":
            return "school_telephone"
        if new_name == "iu_website":
            return None
        if new_name == "iu_name":
            return None
        return new_name

    schools = {}
    sheet = wb.active

    first = True
    for row in sheet.rows:
        if first:
            first = False
            continue

        lea_name = row[0].value
        school_name = row[1].value
        aun = detect_type(row[2].value)
        school_id = detect_type(row[3].value)
        attr = rename_school_fast_facts_attribute(row[4].value)
        value = detect_type(row[5].value)


        if attr is None:
            continue
        if "lea_1" in attr or "lea_2" in attr:
            continue

        if school_id not in schools:
            schools[school_id] = {"school_name": school_name, "lea_name": lea_name, "aun": aun}

        schools[school_id][attr] = value

    return SheetDict(schools, "school_id")

def parse_afr_expenditure(wb, year):
    '''
    Parses AFR Expenditure worksheets.

    <ARGUMENTS>
        * wb [openpyxl Workbook]: The worksheet to be parsed.

    <RETURN>
        * [utils.SheetDict]:
            A SheetDict containing the parsed AFR Expenditure data (Sheet 1/2).

        * [utils.SheetDict]:
            A SheetDict containing the parsed AFR Expenditure per ADM data
            (Sheet 2/2).
    '''

    def get_exp_aun(row, year):
        return row[1].value

    def get_exp_adm_aun(row, year):
        if year == 2018 or year == 2022:
            return row[1].value
        return row[0].value

    def rename_afr_exp_attr(attribute):
        new_name = rename_attribute(attribute)
        if new_name is None:
            return None
        if "actual_instruction_expense" in new_name:
            return None

        search = re.search(r'\d{4}$', new_name)

        if search is None:

            if new_name == "aun":
                return None
            if "cuurent" in new_name:
                return None
            if "current" in new_name:
                return None
            if new_name == "ctgy":
                return None
            if new_name == "cat":
                return None
            if "actual_instruction" in new_name:
                return None

            return new_name

        id = int(search.group(0))
        match id:
            case 1000:
                return "instruction"
            case 1100:
                return "regular_programs"
            case 1200:
                return "gifted_programs"
            case 1300:
                return "vocational_programs"
            case 1400:
                return "other_instructional"
            case 1500:
                return "nonpublic_programs"
            case 1600:
                return "adult_programs"
            case 1700:
                return "secondary_programs"
            case 1800:
                return "pre_k"
            case 2000:
                return "support_services"
            case 2100:
                return "personnel"
            case 2200:
                return "staff"
            case 2300:
                return "administration"
            case 2400:
                return "health"
            case 2500:
                return "business"
            case 2600:
                return "plant_ops"
            case 2700:
                return "transportation"
            case 2800:
                return "central"
            case 2900:
                return "other_support"
            case 3000:
                return "noninstructional_services"
            case 4000:
                return "facai"
            case 5000:
                return "oefu"
            case _:
                print("Invalid ID!")
                print(id)
                print(new_name)
                exit()

    def rename_afr_exp_adm_attr(attribute):
        new_name = rename_attribute(attribute)
        if new_name is None:
            return None
        if new_name == "aun":
            return None
        if new_name == "ctgy":
                return None
        if new_name == "cat":
            return None

        if bool(re.match(r'^\d{4}\-\d{2}\_', new_name)):
            new_name = new_name[8:]

        new_name = new_name.replace("memebership", "membership")
        new_name = new_name.replace("average_daily_membership", "adm")

        if not (new_name == "adm" or new_name == "weighted_adm"):
            return None

        return new_name


    expenditures = parse_standard_sheet(wb.worksheets[0], year, rename_afr_exp_attr, get_exp_aun)
    expenditures_dict =  SheetDict(expenditures, "aun")

    expenditures_adm = parse_standard_sheet(wb.worksheets[1], year, rename_afr_exp_adm_attr, get_exp_adm_aun)
    expenditures_adm_dict =  SheetDict(expenditures_adm, "aun")

    return [expenditures_dict, expenditures_adm_dict]

def parse_afr_revenue(wb, year):
    '''
    Parses AFR Revenue worksheets.

    <ARGUMENTS>
        * wb [openpyxl Workbook]: The worksheet to be parsed.

    <RETURN>
        * [utils.SheetDict]:
            A SheetDict containing the parsed AFR Revenue data (Sheet 1/3).

        * [utils.SheetDict]:
            A SheetDict containing the parsed AFR Reveune Tax Colleted & Eq.
            Millis data (Sheet 2/3).

        * [utils.SheetDict]:
            A SheetDict containing the parsed AFR Reveune per ADM data (Sheet 3/3).
    '''

    def get_rbs_aun(row, year):
        return row[1].value

    def get_rpa_aun(row, year):
        if year == 2019:
            return row[1].value
        return row[0].value

    def get_tcem_aun(row, year):
        return row[0].value

    def rename_rbs_attr(attr):
        new_name = rename_attribute(attr)

        if new_name is None:
            return None
        if "total_local" in new_name:
            return None
        #if "total_revenue" in new_name:
        #    return None
        if "%" in new_name:
            return None

        search = re.search(r'\d{4}', new_name)

        if search is None:

            if new_name == "aun":
                return None
            if "cuurent" in new_name:
                return None
            if "current" in new_name:
                return None
            if new_name == "ctgy":
                return None
            if new_name == "cat":
                return None

            return new_name

        id = int(search.group(0))
        match id:
            case 6111:
                return "local_taxes"
            case 6500:
                return "local_other"
            case 7000:
                return "state_revenue"
            case 8000:
                return "federal_revenue"
            case 9000:
                return "other_revenue"
            case _:
                print("Invalid ID!")
                print(id)
                print(new_name)
                exit()

    def rename_rpa_attr(attr):
        new_name = rename_attribute(attr)

        if new_name is None:
            return None
        if new_name == "aun":
            return None
        if new_name == "cat":
            return None
        if "rank" in new_name:
            return None
        if  "total" in new_name:
            return None
        if "average_daily" in new_name:
            return "adm"
        return new_name

    def rename_tcem_attr(attr):
        new_name = rename_attribute(attr)

        if new_name is None:
            return None
        if new_name == "aun":
            return None
        if "total" in new_name:
            return None
        if "rank" in new_name:
            return None
        if "steb" in new_name:
            return new_name[5:]
        if "equalized" in new_name:
            return new_name[8:]
        if "679" in new_name:
            return new_name[4:]
        return new_name

    if year in [2013, 2017, 2018, 2022]:
        rpa_sheet = wb.worksheets[1]
        tcem_sheet = wb.worksheets[2]
    else:
        rpa_sheet = wb.worksheets[2]
        tcem_sheet = wb.worksheets[1]

    rbs = parse_standard_sheet(wb.worksheets[0], year, rename_rbs_attr, get_rbs_aun)
    rpa = parse_standard_sheet(rpa_sheet, year, rename_rpa_attr, get_rpa_aun)
    tcem = parse_standard_sheet(tcem_sheet, year, rename_tcem_attr, get_tcem_aun)

    rbs_dict = SheetDict(rbs, "aun")
    rpa_dict = SheetDict(rpa, "aun")
    tcem_dict = SheetDict(tcem, "aun")

    return [rbs_dict, rpa_dict, tcem_dict]

def parse_aid_ratio(wb, year):
    '''
    Parses Aid Ratio worksheets.

    <ARGUMENTS>
        * wb [openpyxl Workbook]: The worksheet to be parsed.

    <RETURN>
        * [utils.SheetDict]:
            A SheetDict containing the parsed Aid Ratio LEA data (Sheet 1/2).

        * [utils.SheetDict]:
            A SheetDict containing the parsed Aid Ratio IU data (Sheet 2/2).
    '''

    def rename_aid_ratio_attr(attr):

        new_name = rename_attribute(attr)

        if new_name is None:
            return None

        new_name = new_name.replace("market_value", "mv")
        new_name = new_name.replace("personal_income", "pi")

        if new_name == "h":
            return None
        if new_name == "aun":
            return None
        if new_name == "fiscal_year":
            return None
        if new_name == "career_and_technology_center_(and_participating_sd)":
            return "lea_name"
        if  "filter" in new_name:
             return None
        if "sort" in new_name:
            return None
        if "ctc_aun" in new_name:
            return None
        if "iu_aun" in new_name:
            return None
        if "cyber" in new_name:
            return None
        if "cs_aun" in new_name:
            return None
        if "july_2020" in new_name: # This case and the next only apply to 2020-2021
            return None
        if "percent" in new_name:
            return None
        if "mv_/_pi" in new_name:
            return "mv_pi_aid_ratio"
        if "charter_school" in new_name:
            return "lea_name"

        new_name = new_name.replace("wadm", "weighted_adm")

        if "weighted_adm" in new_name:
            if "mv" in new_name or "pi" in new_name:
                return None
            return "weighted_adm"

        if "ratio" in new_name:
            if "mv" in new_name:
                return "mv_aid_ratio"
            if "pi" in new_name:
                return "pi_aid_ratio"
            print("Invalid Match:")
            print(attr)
            exit()

        if "mv" in new_name:
            return "mv"
        if "pi" in new_name:
            return "pi"

        return new_name

    def rename_aid_ratio_iu_attr(attr):
        new_name = rename_aid_ratio_attr(attr)

        if new_name == "intermediate_unit_(and_participating_sd)":
            return "iu_name"
        if new_name == "lea_name":
            return "iu_name"
        return new_name

    def get_sd_id(row, year):
        if year <= 2016:
            return row[0].value
        return row[1].value

    def get_iu_id(row, year):
        if year <= 2016:
            aun = row[0].value
        else:
            aun = row[1].value

        if str(aun).endswith("000000"):
            return aun
        return None

    def get_ctc_id(row, year):
        if year <= 2016:
            aun = row[0].value
        else:
            aun = row[1].value

        if str(aun).endswith("07") or str(aun).endswith("57"):
            return aun

        return None

    def get_cs_id(row, year):
        if year < 2015:
            return row[1].value
        if year <= 2016:
            return row[0].value
        return row[1].value

    lea = parse_standard_sheet(wb.worksheets[0], year, rename_aid_ratio_attr, get_sd_id)
    iu = parse_standard_sheet(wb.worksheets[1], year, rename_aid_ratio_iu_attr, get_iu_id)
    ctc = parse_standard_sheet(wb.worksheets[2], year, rename_aid_ratio_attr, get_ctc_id)
    cs = parse_standard_sheet(wb.worksheets[3], year, rename_aid_ratio_attr, get_cs_id)

    lea_dict = SheetDict(lea | ctc | cs, "aun")
    iu_dict = SheetDict(iu, "aun")

    return [lea_dict, iu_dict]

def parse_keystone(wb, year):
    '''
    Parses Keystone Exam worksheets.

    <ENTENDED_DESCRIPTION>
    Because the Keystone Exam worksheets follow an unusual file structure
    (with a key composed of multiple attributes), this function cannot rely
    on @parse_standard_sheet(...). Instead, it implements a custom parsing
    algorithm tailored to this unique structure.

    <ARGUMENTS>
        * wb [openpyxl Workbook]: The worksheet to be parsed.

    <RETURN>
        * [utils.SheetDict]: A SheetDict containing the parsed Keystone Exam Data.
    '''

    def get_keystone_school_id(row, year):
        if year == 2015:
            return row[0].value
        if year in [2017, 2018]:
            return row[1].value
        if year in [2016, 2019, 2021, 2022]:
            return row[2].value
        return row[2].value

    def rename_keystone_attr(attr):
        new_name = rename_attribute(attr)
        if new_name is None:
            return None
        if new_name == "year":
            return None
        if new_name == "growth":
            return None
        if "2019" in new_name:
            return None
        if "below" in new_name:
            return "below_basic"
        if "basic" in new_name:
            return "basic"
        if "proficient" in new_name:
            return "proficient"
        if "advanced" in new_name:
            return "advanced"
        if new_name == "all_students":
            return "all"
        if new_name == "economically_disadvantaged":
            return "ed"
        if new_name == "historically_underperforming":
            return "hu"
        if new_name == "student_group_name":
            return "group"
        if new_name == "n_scored" or new_name == "number_scored":
            return "scored"
        if new_name == "m" or new_name == "algebra_i" or new_name == "algebra_1":
            return "algebra"
        if new_name == "e":
            return "literature"
        if new_name == "s":
            return "biology"

        return new_name

    if year == 2015:
        keystone_sheet = trim_sheet(wb.worksheets[0], 7, 0)
    elif year == 2021:
        keystone_sheet = trim_sheet(wb.worksheets[0], 5, 0)
    else:
        keystone_sheet = trim_sheet(wb.worksheets[0], 4, 0)

    attr_idxs = {}
    for col_idx, cell in enumerate(next(keystone_sheet.rows)):
        attr = rename_keystone_attr(cell.value)
        if attr is not None:
            attr_idxs[attr] = col_idx

    school_id_idx = attr_idxs["school_id"]
    subject_idx = attr_idxs["subject"]
    group_idx = attr_idxs["group"]
    rating_idxs = [attr_idxs["advanced"], attr_idxs["proficient"], attr_idxs["basic"], attr_idxs["below_basic"]]

    schools = {}
    for row_idx, row in enumerate(keystone_sheet.rows):
        school_id = detect_type(row[school_id_idx].value)
        group = rename_keystone_attr(row[group_idx].value)
        subject = rename_keystone_attr(row[subject_idx].value)

        if row_idx == 0:
            continue
        if school_id is None or group is None or subject is None:
            print(f'Invalid enrtry. ID: {school_id}, Group: {group}, Subject: {subject}')
            continue
        if detect_type(row[attr_idxs["grade"]].value) != 11:
            continue

        composite_key = str(school_id) + "_" + group + "_" + subject

        if composite_key not in schools:
            schools[composite_key] = {}

        for col_idx, cell in enumerate(row):
            attr = rename_keystone_attr(keystone_sheet.cell(row=1, column=col_idx+1).value)
            value = detect_type(cell.value)

            if attr is None:
                continue
            if attr == "grade":
                continue
            if col_idx in [school_id_idx, subject_idx, group_idx]:
                continue

            schools[composite_key][attr] = value

    #print(schools)
    return SheetDict(schools, "school_id")

def parse_apd(wb):
    '''
    Parses APD worksheets.

    <ENTENDED_DESCRIPTION>
    Because the APD worksheets follow an unusual file structure (with what
    should be columns instead being rows), this function cannot rely on
    @parse_standard_sheet(...). Instead, it implements a custom parsing
    algorithm tailored to this unique structure.

    <ARGUMENTS>
        * wb [openpyxl Workbook]: The worksheet to be parsed.

    <RETURN>
        * [utils.SheetDict]: A SheetDict containing the parsed data.
    '''

    def rename_apd_attr(attr):
        new_name = rename_attribute(attr)
        new_name = new_name.replace("_-_", "_").replace("/", "_").replace(":", "_").replace("-", "_").replace("(", "").replace(")", "")

        return new_name

    schools = {}
    sheet = wb.active

    first = True
    for row in sheet.rows:
        if first:
            first = False
            continue

        lea_name = row[0].value
        school_name = row[1].value
        aun = detect_type(row[2].value)
        school_id = detect_type(row[3].value)
        attr = rename_apd_attr(row[4].value)
        value = detect_type(row[5].value)

        if attr is None:
            continue

        if school_id not in schools:
            schools[school_id] = {"school_name": school_name, "lea_name": lea_name, "aun": aun}

        schools[school_id][attr] = value

    return SheetDict(schools, "school_id")

def parse_cohort(wb, year, fix = False):
    '''
    Parses Cohort worksheets.

    <ARGUMENTS>
        * wb [openpyxl Workbook]: The worksheet to be parsed.

    <RETURN>
        * [utils.SheetDict]:
            A SheetDict containing the parsed Cohort LEA data (Sheet 1/2).

        * [utils.SheetDict]:
            A SheetDict containing the parsed Cohort School data (Sheet 2/2).
    '''

    def get_cohort_lea_aun(row, year):
        aun = row[1].value
        if aun == "STATE":
            return None
        if aun == "'-  1  -" or aun == "-  1  -": # This only occurance I found of this was Cohort_Five_Year_2014-2015
            return None

        return aun

    def get_cohort_school_id(row, year):
        school_id = row[3].value
        if school_id == "STATE":
            return None

        return school_id

    def get_cohort_lea_aun_fix(row, year):
        aun = row[0].value
        if aun == "State":
            return None

    def rename_cohort_attr(attr):
        new_name = rename_attribute(attr)

        if new_name == None:
            return None

        if "black" in new_name:
            return new_name.replace("black", "african_american")
        if "american_indian/alaskan_native" in new_name:
            return new_name.replace("american_indian/alaskan_native", "ai_an")
        if "american_indian/_alaskan_native" in new_name:
            return new_name.replace("american_indian/_alaskan_native", "ai_an")
        if "aian" in new_name:
            return new_name.replace("aian", "ai_an")
        if "native_hawaiian_or_pacific_islander" in new_name:
            return new_name.replace("native_hawaiian_or_pacific_islander", "nh_pi")
        if "multi-racial" in new_name:
            return new_name.replace("multi-racial", "multiracial")
        if "sp_ed_" in new_name:
            return new_name.replace("sp_ed_", "special_ed_")
        if "ell_" in new_name or "el_" in new_name:
            return new_name.replace("ell_", "english_learner_").replace("el_", "english_learner_")
        if "econ_disadv_" in new_name:
            return new_name.replace("econ_disadv_", "economically_disadvantaged_")
        if new_name == "grades" or new_name == "grads":
            return "total_grads"
        if new_name == "cohort_grad_rate" or new_name == "grad_rate" or new_name == "total_grad_rate": # unecessary because it can be calculated
            return None
        if new_name == "lea":
            return "lea_name"
        if new_name == "cohort":
            return "total_cohort"
        if new_name == "school_number":
            return "school_id"
        if new_name == "school":
            return "school_name"

        return new_name

    def rename_cohort_lea_attr(attr):
        new_name = rename_cohort_attr(attr)

        if new_name == "aun":
            return None

        return new_name

    def rename_cohort_lea_attr_fix(attr):
        new_name = rename_cohort_lea_attr(attr)
        if new_name == "lea_type":
            return "aun"
        if new_name == "aun":
            return "lea_name"
        if new_name == "lea":
            return lea_type

        return new_name

    def rename_cohort_school_attr(attr):
        new_name = rename_cohort_attr(attr)

        if new_name == "school_id":
            return None

        return new_name

    if fix: # This is needed because Cohort_Four_Year_2012-2013 misslabled the attributes. This swaps the names/positions so that they match up again
        cohort_lea_sheet = parse_standard_sheet(trim_sheet(wb.worksheets[2], 2, 1), year, rename_cohort_lea_attr_fix, get_cohort_lea_aun_fix)
    else:
        cohort_lea_sheet = parse_standard_sheet(trim_sheet(wb.worksheets[2], 2, 1), year, rename_cohort_lea_attr, get_cohort_lea_aun)

    cohort_sch_sheet = parse_standard_sheet(trim_sheet(wb.worksheets[3], 2, 1), year, rename_cohort_school_attr, get_cohort_school_id)

    cohort_lea_dict = SheetDict(cohort_lea_sheet, "aun")
    cohort_sch_dict = SheetDict(cohort_sch_sheet, "school_id")

    return [cohort_lea_dict, cohort_sch_dict]

def write_dicts(classified_sheet_dicts, subdirectory, CLEAN_DATA_DIRECTORY):
    '''
    Writes the SheetDicts to file.

    <ARGUMENTS>
        * classified_sheet_dicts [Dictionary]:
            A dictionary of SheetDicts, where the key is the type of statistic.

        * subdirectory [String]:
            The subdirectory containing a data file.

        * CLEAN_DATA_DIRECTORY [String]:
            The path to the output directory.
    '''

    Path(CLEAN_DATA_DIRECTORY + "/" + subdirectory).mkdir(parents=True, exist_ok=True)
    for classification, sheet_dicts in classified_sheet_dicts.items():
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        next_key_index = 2
        key_indices = {}

        for year, sheet_dict in dict(sorted(sheet_dicts.items())).items():
            #if not bool(sheet_dict.dict): # Doesn't write an empty dict
            #    continue

            sheet_name = str(year)
            sheet = wb.create_sheet(title=sheet_name)

            rowIdx = 2

            sheet.cell(row=1, column=1).value = sheet_dict.identifier

            for key, record in sheet_dict.dict.items():
                sheet.cell(row=rowIdx, column=1).value = key

                for record_key, attribute in record.items():
                    if record_key not in key_indices:
                        key_indices[record_key] = next_key_index
                        next_key_index = next_key_index + 1


                    sheet.cell(row=1, column = key_indices[record_key]).value = record_key

                    sheet.cell(row=rowIdx, column=key_indices[record_key]).value = attribute
                rowIdx = rowIdx + 1

        wb.save(CLEAN_DATA_DIRECTORY + "/" + subdirectory + "/" + classification + ".xlsx")
        wb.close()