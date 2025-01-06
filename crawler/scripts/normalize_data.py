'''
<FILE>
normalize_data.py


<DESCRIPTION>
The purpose of this script is to normalize the data output from <clean_data.py>, and
prepare it for insertion into the database. It does so through a few important steps:

    * Combine Years:
        Each input file has several sheets, one for each year in which
        the data was recorded. This script combines these sheets into
        one, where the year is part of the primary key.

    * 2NF Sheets:
        Some sheets might have functional dependencies that can be removed
        (e.g. aun -> lea_name). These are pulled out into separate sheets
        (IUs, LEAs, Schools).

    * Label Data with Types:
        Automatically detects what SQLite3 type (TEXT, REAL, INT) each
        attribute is. It then labels each attribute with this type.

<FUNCTIONS>
This script can be run by calling normalize_data.run(<args>). All other functions
in this script should remain private. This section only lists a brief description
of each function. For more comprehensive documentation, see each method directly.

    * run(...):
        Normalizes the output from <clean_data.py>.

    * can_safely_replace(...):
        Detects if a value can safely be clobbered by another.

    * add_to_sheet_dict(...):
        Adds a key/value pair to a SheetDict.

    * add_to_composite_dict(...):
        Adds a key/value pair to a composite dictionary.

    * merge_composite_dicts(...):
        Combines two composite dictionaries into one.

    * parse_standard_wb(...):
        Parses standard workbook to write the data into
        corresponding SheetDict/composite dictionaries.

    * parse_ffs_wb(...):
        Parses a Fast Facts School workbook to write the
        data into corresponding SheetDict/composite dictionaries.

    * parse_keystone_wb(...):
        Parses a Keystone workbook to write the data into
        corresponding SheetDict/composite dictionaries.

    * write_composite_dict(...):
        Writes a composite dictionary into file.

    * write_sheet_dict(...):
        Writes a SheetDict into file.
'''

import openpyxl
import shutil
import os
import re
from xls2xlsx import XLS2XLSX
from scripts.utils import Logger, detect_year, detect_type, detect_db_type, SheetDict


schools = SheetDict({}, "school_id")
leas = SheetDict({}, "aun")
ius = SheetDict({}, "aun")

def run(CLEAN_DATA_DIRECTORY, NORMALIZED_DATA_DIRECTORY, logger):
    '''
     Normalizes the output from <clean_data.py>.

    <EXTENDED_DESCRIPTION>
    This is the only intended entry-point into <normalize_data.py>. This
    script will read in the files from the clean data directory, and write
    the results of this script to the normalized data directory.

    <ARGUMENTS>
        * CLEAN_DATA_DIRECTORY [String]: The path to the input directory.

        * NORMALIZED_DATA_DIRECTORY [String]: The path to the output directory.

        * logger [utils.Logger]: The current Logger instance.
    '''

    logger.indent()

    for subdirectory in os.listdir(CLEAN_DATA_DIRECTORY):
        data_dict = SheetDict({}, "")
        col_types_dict = {}

        new_file = NORMALIZED_DATA_DIRECTORY + "/" + subdirectory + ".xlsx"
        for filename in os.listdir(CLEAN_DATA_DIRECTORY + "/" + subdirectory):
            if "#" in filename:
                continue

            logger.write(f'Processing {subdirectory}/{filename}')
            #if "AFR" not in filename and "IU" not in filename and "Fast_Facts" not in filename and "LEA" not in filename and "Keystone" not in filename:
            #    continue
            #if "Keystone" not in filename:
            #    continue
            if "APD" in filename:
                continue

            file = CLEAN_DATA_DIRECTORY + "/" + subdirectory + "/" + filename

            #print(f'File: {file}')

            wb = openpyxl.open(file)

            if "Fast_Facts_School" in filename:
                [dict, col_types] = parse_ffs_wb(wb, logger)
            elif "Cohort_School" in filename:
                pass
                #[dict, col_types] = parse_cohort_school(wb, logger)
            elif "Keystone" in filename:
                [dict, col_types] = parse_keystone_wb(wb, logger)
            else:
                [dict, col_types] = parse_standard_wb(wb, logger)

            merge_composite_dicts(logger, data_dict, dict)

            col_types_dict.update(col_types)

            wb.close()

        #print(f'Data dict at end: {data_dict}')
        write_composite_dict(data_dict, col_types_dict, new_file)


    write_sheet_dict(schools, NORMALIZED_DATA_DIRECTORY + "/Schools.xlsx")
    write_sheet_dict(leas, NORMALIZED_DATA_DIRECTORY + "/LEAs.xlsx")
    write_sheet_dict(ius, NORMALIZED_DATA_DIRECTORY + "/IUs.xlsx")

    logger.unindent()

#logger = Logger("crawler-logs.txt")
#logger.write("Staring Script...")
#run("./data-organized", "./data-clean", logger)
#logger.write("Done!")


def can_safely_replace(val1, val2):
    '''
    Detects if a value can be safely clobbered by another.

    <EXTENDED_DESCRIPTION>
    Dictionaries can only have one value per key. However, in some cases,
    that value may need to be updated (e.g. a School District changes its
    name). This function detects if this change is a trivial one, such as a
    capitalization change, and thus determines if the value can be safely
    replaced without losing data.

    <ARGUMENTS>
        * val1 [Any]: The original value.

        * val2 [Any]: The replacement value.

    <RETURNS>
        * [Boolean]: If a value can be safely clobbered by another.
    '''
    if val1 == val2:
        return True

    if not (type(val1) == str and type(val2) == str):
        return False

    val1 = val1.replace(" CHS", "CS")
    val2 = val2.replace(" CHS", "CS")

    removal_list = [" ", ".", ",", "-", "&"]
    replacement_list = [("charterschool", "cs"), ("charterhighschool", "cs"), ("charterscholars", "cs"), ("saint", "st"), ("road", "rd"), ("drive", "dr"), ("careerandtechnicalcenter", "ctc"), ("careertechnicalcenter", "ctc"), ("schooldistrict", "sd"), ("charterelementaryschool", "ces")]

    val1 = val1.lower()
    val2 = val2.lower()

    for removal in removal_list:
        val1 = val1.replace(removal, "")
        val2 = val2.replace(removal, "")

    for replacement in replacement_list:
        val1 = val1.replace(replacement[0], replacement[1])
        val2 = val2.replace(replacement[0], replacement[1])

    return val1 == val2


def add_to_sheet_dict(logger, sheet_dict, id, attribute, value):
    '''
    Adds a key/value pair to a SheetDict.

    <EXTENDED_DESCRIPTION>
    See <utils.py> for a description of SheetDicts.

    Adds a given attribute and value to a given SheetDict.

    <ARGUMENTS>
        * logger [utils.Logger]:
            The current Logger instance.

        * sheet_dict [utils.SheetDict]:
            The SheetDict in which the key/value pair should be written to.

        * id [Integer]:
            The record that should be written to.

        * attribute [String]:
            The name of the attribute being written to in the record.

        * value [Any]:
            The value being written into the record.
    '''

    if id is None:
        return
    if value is None:
        return

    dict = sheet_dict.dict
    if id not in dict:
        dict[id] = {}

    if attribute in dict[id]:
        old_val = dict[id][attribute]
        if not can_safely_replace(old_val, value):
            logger.write(f'Clobbering {attribute} in dict[{id}]. Replacing {old_val} with {value}');

    dict[id][attribute] = value

def add_to_composite_dict(logger, composite_dict, id, year, attribute, value):
    '''
    Adds a key/value pair to a composite dict.

    <EXTENDED_DESCRIPTION>
    A composite dictionary is similar to a SheetDict (See <utils.py>), but having
    a record where the key is composed both of an id and the year in which the
    data was recorded. E.g., each school has multiple attribute/value pairs for
    each calendar year.

    <ARGUMENTS>
        * logger [utils.Logger]:
            The current Logger instance.

        * composite_dict [Dictionary]:
            The composite dictionary in which the data should be written to.

        * id [Integer]:
            The record that should be written to in the composite dictionary.

        * year [Integer]:
            The year in which the data was recorded.

        * attribute [String]:
            The name of the attribute being written to in the record.

        * value [Any]:
            The value being written into the record.
    '''

    if value is None:
        return
    if id is None:
        return

    if year not in composite_dict:
        composite_dict[year] = {}
    if id not in composite_dict[year]:
        composite_dict[year][id] = {}

    if attribute in composite_dict[year][id]:
        old_val = composite_dict[year][id][attribute]
        if not can_safely_replace(old_val, value):
            logger.write(f'Clobbering {attribute} in composite_dict[{year}][{id}]. Replacing {old_val} with {value}');

    composite_dict[year][id][attribute] = value

def merge_composite_dicts(logger, dest, source):
    '''
    Merges two composite dictionaries.

    <EXTENDED_DESCRIPTION>
    See @add_to_composite_dict(...) for a description of composite dictionaries.

    Combines two composite dictionaries into one. This is useful because multiple
    files might contain similar data, so it might make sense to combine the
    data points into a singular composite dictionary.

    <ARGUMENTS>
        * logger [utils.Logger]:
            The current Logger instance.

        * dest [Dictionary]:
            The composite dictionary which is being written to.

        * source [Dictionary]:
            The composite dictionary which is being read from.
    '''

    dest_dict = dest.dict
    source_dict = source.dict

    for year, year_dict in source_dict.items():
        if year not in dest_dict:
            dest_dict[year] = {}

        for id, id_dict in year_dict.items():
            if id not in dest_dict[year]:
                dest_dict[year][id] = {}

            dest_dict[year][id].update(id_dict)
            #print(f'Yearss: {dest_dict[year]}')

        #if id not in dest_dict[year].keys():
        #    dest_dict[year][id] = {}

    if not (dest.identifier == "" or dest.identifier == source.identifier):
        logger.warn(f'Dicts have different identifier. Source: {source.identifier}, Dest: {dest.identifier}')

    dest.identifier = source.identifier

def parse_standard_wb(wb, logger):
    '''
    Parses standard workbook to write the data into corresponding SheetDict/
    composite dictionaries.

    <EXTENDED_DESCRIPTION>
    See @add_to_composite_dict(...) for a description of composite dictionaries.
    See <utils.py> for a description of SheetDicts.

    Parses a standard workbook. A "standard" workbook assumes that the
    data describes a LEA, and the AUN is present in the first column.

    <ARGUMENTS>
        * wb [openpyxl.Workbook]: The workbook to be parsed.

        * logger [util.Logger]: The current Logger instance.

    <RETURN>
        * [util.SheetDict]:
            A SheetDict with the parsed data.

        * [Dictionary]:
            A dictionary with the SQLite3 types that best describes
            each attribute.
    '''

    data_dict = {}
    col_types = {}

    logger.indent()

    for sheet in wb.worksheets:
        year = detect_type(sheet.title)

        for col_idx, col in enumerate(sheet.iter_cols()):
            attr = col[0].value
            if attr is None:
                continue

            col_types[attr] = detect_db_type(col)

            if col_idx == 0:
                continue

            is_lea_attr = attr in ["lea_type", "lea_name", "county", "lea_address_street", "lea_address_city", "lea_address_state", "lea_address_zip", "lea_website", "lea_telephone"]
            is_iu_attr = attr in ["iu_name"]
            #is_school_attr = attr in ["school_name", "school_address_street", "school_address_state", "school_address_zip", "school_website", "school_telephone"]

            for row_idx, cell in enumerate(col):
                if row_idx == 0:
                    continue

                aun = sheet.cell(row=row_idx+1, column=1).value
                value = cell.value

                #if attr == "lea_type":
                #    print("SDLFJKJSLKDFJLSKDJF")
                #    exit()
                if not isinstance(aun, int):
                    print(year)
                    exit()

                if is_lea_attr:
                    add_to_sheet_dict(logger, leas, aun, attr, value)
                elif is_iu_attr:
                    if "000000" in str(aun):
                        add_to_sheet_dict(logger, ius, aun, attr, value)
                    #logger.warn(f'Cannot add to iu dict. Attr: {attr}, Val: {value}' )
                else:
                    add_to_composite_dict(logger, data_dict, aun, year, attr, value)


    logger.unindent()
    return (SheetDict(data_dict, "aun"), col_types)

def parse_ffs_wb(wb, logger):
    '''
    Parses a Fast Facts School workbook to write the data into corresponding
    SheetDict/composite dictionaries.

    <EXTENDED_DESCRIPTION>
    See @add_to_composite_dict(...) for a description of composite dictionaries.
    See <utils.py> for a description of SheetDicts.

    Parses a Fast Facts School workbook.

    <ARGUMENTS>
        * wb [openpyxl.Workbook]: The workbook to be parsed.

        * logger [util.Logger]: The current Logger instance.

    <RETURN>
        * [util.SheetDict]:
            A SheetDict with the parsed data.

        * [Dictionary]:
            A dictionary with the SQLite3 types that best describes
            each attribute.
    '''

    data_dict = {}
    col_types = {}

    logger.indent()

    for sheet in wb.worksheets:
        year = detect_type(sheet.title)

        for col_idx, col in enumerate(sheet.iter_cols()):
            attr = col[0].value
            if attr is None:
                continue

            col_types[attr] = detect_db_type(col)

        #print(col_types)

        for row_idx, row in enumerate(sheet.iter_rows()):
            if row_idx == 0:
                continue

            for col_idx, cell in enumerate(row):
                attr = detect_type(sheet.cell(row=1, column=col_idx+1).value)
                school_id = detect_type(sheet.cell(row=row_idx+1, column=1).value)
                aun = detect_type(sheet.cell(row=row_idx+1, column=4).value)
                value = detect_type(cell.value)

                if attr is None:
                    continue
                if attr == "school_id":
                    continue

                is_lea_attr = attr in ["lea_name", "county", "lea_name", "lea_address_street", "lea_address_city", "lea_address_state", "lea_address_zip", "lea_website", "lea_telephone"]
                is_iu_attr = attr in ["iu_name"]
                is_school_attr = attr in ["school_name", "aun", "school_address_street", "school_address_city", "school_address_state", "school_address_zip", "school_website", "school_telephone"]

                if is_lea_attr:
                    pass
                    #add_to_sheet_dict(logger, leas, aun, attr, value)
                    #logger.warn(f'Cannot add to lea dict. Attr: {attr}, Val: {value}')
                elif is_iu_attr:
                    logger.warn(f'Cannot add to iu dict. Attr: {attr}, Val: {value}')
                elif is_school_attr:
                    add_to_sheet_dict(logger, schools, school_id, attr, value)
                else:
                    add_to_composite_dict(logger, data_dict, school_id, year, attr, value)

                # This bit of code is kinda bad. It is meant to handle CTCs, which are LEAs but not SDs
                school_name = sheet.cell(row=row_idx+1, column=2).value
                lea_name = sheet.cell(row=row_idx+1, column=3).value
                if school_name == lea_name and is_school_attr and attr != "aun":
                    #logger.write(f'fast fact attr: school_id: {school_id}, year: {year}, attr: {attr}. aun: {aun}')
                    add_to_sheet_dict(logger, leas, aun, attr.replace("school_", "lea_"), value)

    logger.unindent()
    return (SheetDict(data_dict, "school_id"), col_types)

def parse_keystone_wb(wb, logger):
    '''
    Parses a Keystone workbook to write the data into corresponding.
    SheetDict/composite dictionaries.

    <EXTENDED_DESCRIPTION>
    See @add_to_composite_dict(...) for a description of composite dictionaries.
    See <utils.py> for a description of SheetDicts.

    <ARGUMENTS>
        * wb [openpyxl.Workbook]: The workbook to be parsed.

        * logger [util.Logger]: The current Logger instance.

    <RETURN>
        * [util.SheetDict]:
            A SheetDict with the parsed data.

        * [Dictionary]:
            A dictionary with the SQLite3 types that best describes
            each attribute.
    '''
    data_dict = {}
    col_types = {}

    logger.indent()

    for sheet in wb.worksheets:
        year = detect_type(sheet.title)

        for col_idx, col in enumerate(sheet.iter_cols()):
            attr = col[0].value
            if attr is None:
                continue

            col_types[attr] = detect_db_type(col)

        #print(col_types)

        for row_idx, row in enumerate(sheet.iter_rows()):
            if row_idx == 0:
                continue

            for col_idx, cell in enumerate(row):
                attr = detect_type(sheet.cell(row=1, column=col_idx+1).value)
                school_id = detect_type(sheet.cell(row=row_idx+1, column=1).value)
                aun = detect_type(sheet.cell(row=row_idx+1, column=29).value)
                value = detect_type(cell.value)

                if attr is None:
                    continue
                if attr == "school_id":
                    continue

                is_lea_attr = attr in ["lea_name", "aun", "county", "lea_name", "lea_address_street", "lea_address_city", "lea_address_state", "lea_address_zip", "lea_website", "lea_telephone"]
                is_iu_attr = attr in ["iu_name"]
                is_school_attr = attr in ["school_name", "aun", "school_address_street", "school_address_city", "school_address_state", "school_address_zip", "school_website", "school_telephone"]

                if is_lea_attr:
                    pass
                    #add_to_sheet_dict(logger, leas, aun, attr, value)
                    #logger.warn(f'Cannot add to lea dict. Attr: {attr}, Val: {value}')
                elif is_iu_attr:
                    logger.warn(f'Cannot add to iu dict. Attr: {attr}, Val: {value}')
                elif is_school_attr and (attr != "aun" or aun is not None):
                    #if aun is None:
                    #    print(attr)
                    add_to_sheet_dict(logger, schools, int(school_id.split("_")[0]), attr, value)
                else:
                    add_to_composite_dict(logger, data_dict, school_id, year, attr, value)

                # This bit of code is kinda bad. It is meant to handle CTCs, which are LEAs but not SDs
                school_name = sheet.cell(row=row_idx+1, column=2).value
                lea_name = sheet.cell(row=row_idx+1, column=3).value
                if school_name == lea_name and is_school_attr and attr != "aun":
                    #logger.write(f'fast fact attr: school_id: {school_id}, year: {year}, attr: {attr}. aun: {aun}')
                    add_to_sheet_dict(logger, leas, aun, attr.replace("school_", "lea_"), value)

    logger.unindent()
    return (SheetDict(data_dict, "school_id"), col_types)

#def parse_cohort_school(wb, logger):
#    pass

def write_composite_dict(composite_dict, col_types, filename):
    '''
    Writes a composite dictionary into file.

    <EXTENDED_DESCRIPTION>
    See @add_to_composite_dict(...) for a description of composite dictionaries.

    <ARGUMENTS>
        * composite_dict [Dictionary]: The composite dictionary to be writted to file.

        * col_types [Dictionary]: The SQLite3 data types for each column.

        * filename [String]: The name of the file to be written to.
    '''

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=1).value = composite_dict.identifier + " PK_INTEGER"
    sheet.cell(row=1, column=2).value = "year PK_INTEGER"

    next_key_index = 3
    key_indices = {}

    rowIdx = 2

    for year, year_dict in composite_dict.dict.items():
        for id, record in year_dict.items():
            for record_key, attribute in record.items():
                sheet.cell(row=rowIdx, column=1).value = id
                sheet.cell(row=rowIdx, column=2).value = year

                if record_key not in key_indices:
                    key_indices[record_key] = next_key_index
                    next_key_index = next_key_index + 1

                if record_key is not None:
                    sheet.cell(row=1, column = key_indices[record_key]).value = record_key + " " + col_types[record_key]
                    #print(f'Record Key: {record_key} + {record_key in col_types}')
                sheet.cell(row=rowIdx, column=key_indices[record_key]).value = attribute
            rowIdx = rowIdx + 1

    wb.save(filename)
    wb.close()

def write_sheet_dict(sheet_dict, filename):
    '''
    Writes a SheetDict into file

    <EXTENDED_DESCRIPTION>
    See <utils.py> for a description of SheetDicts.

    <ARGUMENTS>
        * sheet_dict [utils.SheetDict]: The SheetDict to be writted to file.
        * filename [String]: The name of the file to be written to.
    '''

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = sheet_dict.identifier + " PK_INTEGER"

    next_key_index = 2
    key_indices = {}

    rowIdx = 2

    for id, record in sheet_dict.dict.items():
        for record_key, attribute in record.items():
            sheet.cell(row=rowIdx, column=1).value = id

            if "aun" in record_key or "zip" in record_key or "phone" in record_key:
                key_type = " INTEGER"
            else:
                key_type = " TEXT"

            if record_key not in key_indices:
                key_indices[record_key] = next_key_index
                next_key_index = next_key_index + 1

            sheet.cell(row=1, column = key_indices[record_key]).value = record_key + key_type
            sheet.cell(row=rowIdx, column=key_indices[record_key]).value = attribute
        rowIdx = rowIdx + 1


    wb.save(filename)
    wb.close()